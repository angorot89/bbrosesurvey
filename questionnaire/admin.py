from django.contrib import admin
from django.http import HttpResponse
from django.urls import path, reverse
from django.utils.html import format_html, format_html_join
from django.shortcuts import get_object_or_404
from .models import Employee
import json, csv, io

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def export_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="bbrose_employees.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'Employee ID','Full Name','English Name','Phone','Email','Role','ID Type','ID Number',
        'Language','Created At','Consent',
        'Role & Responsibilities','Workflow & Efficiency','Pain Points',
        'Communication','Tools & Technology','Ideas & Suggestions'
    ])
    for e in queryset:
        writer.writerow([
            e.employee_id, e.full_name, e.english_name, e.phone, e.email, e.role, e.id_type, e.id_number,
            e.language, e.created_at.strftime('%Y-%m-%d %H:%M'),
            'Yes' if e.consent_given else 'No',
            json.dumps(e.role_responsibilities, ensure_ascii=False),
            json.dumps(e.workflow_efficiency, ensure_ascii=False),
            json.dumps(e.pain_points, ensure_ascii=False),
            json.dumps(e.communication, ensure_ascii=False),
            json.dumps(e.tools_technology, ensure_ascii=False),
            json.dumps(e.ideas_suggestions, ensure_ascii=False),
        ])
    return response
export_csv.short_description = "Export selected as CSV"


def export_excel(modeladmin, request, queryset):
    if not HAS_OPENPYXL:
        return HttpResponse("openpyxl not installed", status=500)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    headers = [
        'Employee ID','Full Name','English Name','Phone','Email','Role','ID Type','ID Number',
        'Language','Created At','Consent',
        'Role & Responsibilities','Workflow & Efficiency','Pain Points',
        'Communication','Tools & Technology','Ideas & Suggestions'
    ]
    ws.append(headers)
    for e in queryset:
        ws.append([
            e.employee_id, e.full_name, e.english_name, e.phone, e.email, e.role, e.id_type, e.id_number,
            e.language, e.created_at.strftime('%Y-%m-%d %H:%M'),
            'Yes' if e.consent_given else 'No',
            json.dumps(e.role_responsibilities, ensure_ascii=False),
            json.dumps(e.workflow_efficiency, ensure_ascii=False),
            json.dumps(e.pain_points, ensure_ascii=False),
            json.dumps(e.communication, ensure_ascii=False),
            json.dumps(e.tools_technology, ensure_ascii=False),
            json.dumps(e.ideas_suggestions, ensure_ascii=False),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="bbrose_employees.xlsx"'
    return response
export_excel.short_description = "Export selected as Excel"


def _build_personal_info_workbook(queryset, sheet_title="Personal Info"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    headers = [
        'Employee ID', 'Full Name', 'English Name', 'Phone', 'Email', 'Role',
        'ID Type', 'ID Number', 'Language', 'Consent', 'Created At'
    ]
    ws.append(headers)
    for e in queryset:
        ws.append([
            e.employee_id,
            e.full_name,
            e.english_name,
            e.phone,
            e.email,
            e.role,
            e.id_type,
            e.id_number,
            e.language,
            'Yes' if e.consent_given else 'No',
            e.created_at.strftime('%Y-%m-%d %H:%M'),
        ])
    for idx, width in enumerate([18, 24, 22, 18, 28, 24, 14, 22, 12, 10, 20], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    return wb


def export_selected_personal_info_excel(modeladmin, request, queryset):
    if not HAS_OPENPYXL:
        return HttpResponse("openpyxl not installed", status=500)
    wb = _build_personal_info_workbook(queryset, "Selected Personal Info")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="bbrose_selected_personal_info.xlsx"'
    return response
export_selected_personal_info_excel.short_description = "Export selected personal info as Excel"


def export_all_personal_info_excel(modeladmin, request, queryset):
    if not HAS_OPENPYXL:
        return HttpResponse("openpyxl not installed", status=500)
    wb = _build_personal_info_workbook(Employee.objects.all().order_by('-created_at'), "All Personal Info")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="bbrose_all_personal_info.xlsx"'
    return response
export_all_personal_info_excel.short_description = "Export all users personal info as Excel"


@admin.register(Employee)
class EmployeeAdmin(admin.ModelAdmin):
    list_display = ['employee_id', 'full_name', 'role', 'id_type', 'id_number', 'email', 'language', 'created_at', 'photo_preview']
    list_filter = ['role', 'id_type', 'language', 'created_at', 'consent_given']
    search_fields = ['full_name', 'english_name', 'email', 'employee_id', 'role', 'id_number']
    readonly_fields = ['employee_id', 'created_at', 'updated_at', 'photo_preview_large', 'personal_info_excel_download', 'id_card_preview', 'qr_preview',
                       'display_role_responsibilities', 'display_workflow_efficiency',
                       'display_pain_points', 'display_communication',
                       'display_tools_technology', 'display_ideas_suggestions']
    actions = [export_csv, export_excel, export_selected_personal_info_excel, export_all_personal_info_excel]

    fieldsets = (
        ('Employee Info', {
            'fields': ('employee_id', 'created_at', 'updated_at', 'language')
        }),
        ('Personal Information', {
            'fields': ('full_name', 'english_name', 'phone', 'email', 'role', 'id_type', 'id_number', 'photo', 'photo_preview_large', 'personal_info_excel_download')
        }),
        ('Work Responses', {
            'fields': (
                'display_role_responsibilities',
                'display_workflow_efficiency',
                'display_pain_points',
                'display_communication',
                'display_tools_technology',
                'display_ideas_suggestions',
            ),
            'classes': ('wide',),
        }),
        ('Consent & Generated Assets', {
            'fields': ('consent_given', 'qr_code', 'qr_preview', 'id_card', 'id_card_preview')
        }),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                '<int:employee_id>/download-personal-info/',
                self.admin_site.admin_view(self.download_personal_info_excel),
                name='questionnaire_employee_download_personal_info',
            ),
        ]
        return custom_urls + urls

    def download_personal_info_excel(self, request, employee_id):
        if not HAS_OPENPYXL:
            return HttpResponse("openpyxl not installed", status=500)
        employee = get_object_or_404(Employee, pk=employee_id)
        wb = _build_personal_info_workbook([employee], f"{employee.employee_id}")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{employee.employee_id}_personal_info.xlsx"'
        return response

    def photo_preview(self, obj):
        if obj.photo:
            return format_html('<img src="{}" style="width:40px;height:40px;border-radius:50%;object-fit:cover;" />', obj.photo.url)
        return "—"
    photo_preview.short_description = "Photo"

    def photo_preview_large(self, obj):
        if obj.photo:
            return format_html('<img src="{}" style="width:120px;height:120px;border-radius:12px;object-fit:cover;" />', obj.photo.url)
        return "No photo uploaded"
    photo_preview_large.short_description = "Photo Preview"

    def personal_info_excel_download(self, obj):
        if not obj.pk:
            return "Save the employee first."
        url = reverse('admin:questionnaire_employee_download_personal_info', args=[obj.pk])
        return format_html(
            '<a class="button" href="{}" style="padding:8px 14px;border-radius:6px;background:#0e7490;color:#fff;text-decoration:none;">Download Personal Info Excel</a>',
            url
        )
    personal_info_excel_download.short_description = "Personal Info Export"

    def id_card_preview(self, obj):
        if obj.id_card:
            return format_html('<img src="{}" style="max-width:400px;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,0.15);" />', obj.id_card.url)
        return "Not generated"
    id_card_preview.short_description = "ID Card"

    def qr_preview(self, obj):
        if obj.qr_code:
            return format_html('<img src="{}" style="width:100px;" />', obj.qr_code.url)
        return "Not generated"
    qr_preview.short_description = "QR Code"

    def _format_json(self, title, data):
        if not data:
            return format_html(
                '<div style="max-width:980px;padding:20px 22px;border-radius:18px;'
                'background:linear-gradient(180deg,#0f172a 0%,#172554 100%);'
                'border:1px solid rgba(148,163,184,0.25);box-shadow:0 12px 28px rgba(15,23,42,0.12);'
                'color:#e2e8f0;text-align:left;">'
                '<div style="font-size:1.05rem;font-weight:700;color:#f8fafc;margin-bottom:8px;">{}</div>'
                '<div style="font-size:0.95rem;line-height:1.7;color:#cbd5e1;">No responses provided.</div>'
                '</div>',
                title,
            )

        items_html = format_html_join(
            '',
            '<div style="padding:18px 20px;border-radius:14px;'
            'background:rgba(255,255,255,0.04);border:1px solid rgba(148,163,184,0.18);'
            'margin-bottom:14px;text-align:left;">'
            '<div style="font-size:1rem;font-weight:700;line-height:1.6;color:#f8fafc;'
            'margin-bottom:8px;">{}</div>'
            '<div style="font-size:0.97rem;line-height:1.75;color:#dbeafe;white-space:pre-wrap;">{}</div>'
            '</div>',
            ((question, answer or "—") for question, answer in data.items())
        )

        return format_html(
            '<details open style="max-width:980px;margin:8px 0 18px;">'
            '<summary style="cursor:pointer;list-style:none;outline:none;">'
            '<div style="padding:18px 22px;border-radius:18px 18px 0 0;'
            'background:linear-gradient(180deg,#0f172a 0%,#172554 100%);'
            'border:1px solid rgba(148,163,184,0.25);border-bottom:none;'
            'box-shadow:0 12px 28px rgba(15,23,42,0.12);">'
            '<div style="font-size:1.08rem;font-weight:800;letter-spacing:0.01em;color:#ffffff;">{}</div>'
            '<div style="font-size:0.84rem;color:#bfdbfe;margin-top:6px;">'
            'Question and answer pairs are shown below for quick review.</div>'
            '</div>'
            '</summary>'
            '<div style="padding:18px 18px 6px;border:1px solid rgba(148,163,184,0.25);'
            'border-top:none;border-radius:0 0 18px 18px;'
            'background:linear-gradient(180deg,#172554 0%,#0f172a 100%);">'
            '{}'
            '</div>'
            '</details>',
            title,
            items_html,
        )

    def display_role_responsibilities(self, obj):
        return self._format_json("Role & Responsibilities", obj.role_responsibilities)
    display_role_responsibilities.short_description = "Role & Responsibilities"

    def display_workflow_efficiency(self, obj):
        return self._format_json("Workflow & Efficiency", obj.workflow_efficiency)
    display_workflow_efficiency.short_description = "Workflow & Efficiency"

    def display_pain_points(self, obj):
        return self._format_json("Pain Points", obj.pain_points)
    display_pain_points.short_description = "Pain Points"

    def display_communication(self, obj):
        return self._format_json("Communication", obj.communication)
    display_communication.short_description = "Communication"

    def display_tools_technology(self, obj):
        return self._format_json("Tools & Technology", obj.tools_technology)
    display_tools_technology.short_description = "Tools & Technology"

    def display_ideas_suggestions(self, obj):
        return self._format_json("Ideas & Suggestions", obj.ideas_suggestions)
    display_ideas_suggestions.short_description = "Ideas & Suggestions"
